from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.models.reservation import Reservation, ReservationStatus, PaymentStatus
from app.db.session import get_db
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from datetime import date, timedelta, datetime
from ...models.device import Device
from ...models.reservation import Reservation, ReservationStatus
from urllib.parse import quote
import pandas as pd

router = APIRouter(prefix="/reports", tags=["reports"])

@router.get("/summary")
def report_summary(db: Session = Depends(get_db)):
    """
    报表汇总统计
    """

    total = db.query(func.count(Reservation.id)).scalar() or 0
    completed = (
        db.query(func.count(Reservation.id))
        .filter(Reservation.status == ReservationStatus.COMPLETED)
        .scalar()
        or 0
    )
    cancelled = (
        db.query(func.count(Reservation.id))
        .filter(Reservation.status == ReservationStatus.CANCELLED)
        .scalar()
        or 0
    )
    borrowed = (
        db.query(func.count(Reservation.id))
        .filter(Reservation.status == ReservationStatus.BORROWED)
        .scalar()
        or 0
    )
    total_payment = (
        db.query(func.coalesce(func.sum(Reservation.payment_amount), 0))
        .filter(Reservation.payment_status == PaymentStatus.PAID)
        .scalar()
        or 0.0
    )

    return {
        "total_reservations": total,
        "completed": completed,
        "cancelled": cancelled,
        "borrowed": borrowed,
        "total_payment": float(total_payment),
    }

@router.get("/weekly")
def report_weekly(db: Session = Depends(get_db)):
    start = datetime.now() - timedelta(days=7)

    count = (
        db.query(func.count(Reservation.id))
        .filter(Reservation.start_time >= start)
        .scalar()
        or 0
    )

    return {
        "type": "weekly",
        "start_time": start,
        "total_reservations": count,
    }

@router.get("/monthly")
def report_monthly(db: Session = Depends(get_db)):
    start = datetime.now().replace(day=1)

    count = (
        db.query(func.count(Reservation.id))
        .filter(Reservation.start_time >= start)
        .scalar()
        or 0
    )

    return {
        "type": "monthly",
        "start_time": start,
        "total_reservations": count,
    }

@router.get("/yearly")
def report_yearly(db: Session = Depends(get_db)):
    start = datetime.now().replace(month=1, day=1)

    count = (
        db.query(func.count(Reservation.id))
        .filter(Reservation.start_time >= start)
        .scalar()
        or 0
    )

    return {
        "type": "yearly",
        "start_time": start,
        "total_reservations": count,
    }
@router.get("/weekly/excel", summary="导出周报 Excel")
def export_weekly_report_excel(db: Session = Depends(get_db)):
    # 1. 计算本周时间范围
    today = date.today()
    start_date = today - timedelta(days=today.weekday())
    end_date = start_date + timedelta(days=6)

    # 2. 查询预约数据
    reservations = (
        db.query(Reservation)
        .filter(Reservation.created_at >= start_date)
        .filter(Reservation.created_at <= end_date)
        .all()
    )

    # 3. 创建 Excel（内存中）
    wb = Workbook()
    ws = wb.active
    ws.title = "周报"

    # 表头
    ws.append([
        "预约ID",
        "设备ID",
        "用户ID",
        "开始时间",
        "结束时间",
        "状态",
        "支付状态",
        "金额"
    ])

    # 数据行
    for r in reservations:
        ws.append([
            r.id,
            r.device_id,
            r.user_id,
            r.start_time,
            r.end_time,
            r.status,
            r.payment_status,
            r.payment_amount,
        ])

    # 4. 写入 BytesIO（这就是 output）
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 5. 下载响应
    filename = f"weekly_report_{today}.xlsx"
    filename_cn = f"实验设备周报_{today}.xlsx"

    headers = {
        "Content-Disposition": (
            f"attachment; filename={filename}; "
            f"filename*=UTF-8''{quote(filename_cn)}"
        )
    }

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@router.get("/yearly/excel", summary="导出年报 Excel")
def export_yearly_report_excel(db: Session = Depends(get_db)):
    # 1. 计算本年时间范围
    today = date.today()
    start_date = today.replace(month=1, day=1)
    end_date = today.replace(month=12, day=31)

    # 2. 查询预约数据
    reservations = (
        db.query(Reservation)
        .filter(Reservation.created_at >= start_date)
        .filter(Reservation.created_at <= end_date)
        .all()
    )

    # 3. 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "年报"

    # 表头
    ws.append([
        "预约ID",
        "设备ID",
        "用户ID",
        "开始时间",
        "结束时间",
        "状态",
        "支付状态",
        "金额"
    ])

    # 数据行
    for r in reservations:
        ws.append([
            r.id,
            r.device_id,
            r.user_id,
            r.start_time,
            r.end_time,
            r.status,
            r.payment_status,
            r.payment_amount,
        ])

    # 4. 输出
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename_cn = f"实验设备年报_{today}.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename_cn)}"
    }

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
@router.get("/monthly/excel", summary="导出月报 Excel")
def export_monthly_report_excel(db: Session = Depends(get_db)):
    # 1. 计算本月时间范围
    today = date.today()
    start_date = today.replace(day=1)
    # 月末
    if today.month == 12:
        end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

    # 2. 查询预约数据
    reservations = (
        db.query(Reservation)
        .filter(Reservation.created_at >= start_date)
        .filter(Reservation.created_at <= end_date)
        .all()
    )

    # 3. 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "月报"

    # 表头
    ws.append([
        "预约ID",
        "设备ID",
        "用户ID",
        "开始时间",
        "结束时间",
        "状态",
        "支付状态",
        "金额"
    ])

    # 数据行
    for r in reservations:
        ws.append([
            r.id,
            r.device_id,
            r.user_id,
            r.start_time,
            r.end_time,
            r.status,
            r.payment_status,
            r.payment_amount,
        ])

    # 4. 输出
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename_cn = f"实验设备月报_{today}.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename_cn)}"
    }

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
